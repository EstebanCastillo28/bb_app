import os
import io
import base64
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = "vanti_sales_ultra_secure_key"

# Configuración de base de datos dinámica (Soporta Local SQLite y Vercel Postgres)
DATABASE_URL = os.getenv("POSTGRES_URL") or os.getenv("DATABASE_URL")
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL or "sqlite:///sistema_ventas.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# ──────────────────────────────────────────────
#  MODELOS DE LA BASE DE DATOS
# ──────────────────────────────────────────────

class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False) # 'ADMIN' o 'VENDEDOR'
    ventas = db.relationship('Venta', backref='vendedor', lazy=True)
    pagos = db.relationship('Pago', backref='vendedor', lazy=True)

class Venta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Datos del Cliente
    nombre_cliente = db.Column(db.String(150), nullable=False)
    tipo_doc = db.Column(db.String(50), nullable=False)
    nro_doc = db.Column(db.String(50), nullable=False)
    correo = db.Column(db.String(120))
    telefono = db.Column(db.String(50))
    direccion = db.Column(db.String(200))
    ciudad = db.Column(db.String(100))
    
    # Datos Financieros
    monto_financiado = db.Column(db.Float, nullable=False)
    cantidad_cuotas = db.Column(db.Integer, nullable=False)
    producto_financiado = db.Column(db.String(150), nullable=False)
    observaciones = db.Column(db.Text)
    
    # Archivos adjuntos en Base64
    foto_documento = db.Column(db.Text)
    acta_entrega = db.Column(db.Text)
    foto_cliente_producto = db.Column(db.Text)
    captura_aprobacion = db.Column(db.Text)
    anexos_adicionales = db.Column(db.Text)
    
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)

class Pago(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    monto = db.Column(db.Float, nullable=False)
    observaciones = db.Column(db.Text)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)


# Crear las tablas e inicializar un administrador por defecto si la BD está vacía
with app.app_context():
    db.create_all()
    if not Usuario.query.filter_by(username="admin").first():
        hashed_pw = generate_password_hash("admin123")
        admin_default = Usuario(username="admin", password=hashed_pw, role="ADMIN")
        db.session.add(admin_default)
        db.session.commit()

# ──────────────────────────────────────────────
#  CONTROLADORES Y RUTAS
# ──────────────────────────────────────────────

def procesar_archivo(file_field):
    file = request.files.get(file_field)
    if file and file.filename != '':
        return base64.b64encode(file.read()).decode('utf-8')
    return ""

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username").strip()
        password = request.form.get("password").strip()
        user = Usuario.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password, password):
            session["user_id"] = user.id
            session["username"] = user.username
            session["role"] = user.role
            return redirect(url_for("dashboard"))
        else:
            flash("Credenciales incorrectas.", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/dashboard")
def dashboard():
    if "user_id" not in session: return redirect(url_for("login"))
    
    vendedores = Usuario.query.filter_by(role="VENDEDOR").all() if session["role"] == "ADMIN" else []
    
    # Filtrado por Admin
    filtro_vendedor = request.args.get("vendedor_id")
    
    if session["role"] == "ADMIN":
        if filtro_vendedor:
            ventas_query = Venta.query.filter_by(usuario_id=filtro_vendedor)
            pagos_query = Pago.query.filter_by(usuario_id=filtro_vendedor)
        else:
            ventas_query = Venta.query
            pagos_query = Pago.query
    else:
        # El vendedor solo ve lo suyo
        ventas_query = Venta.query.filter_by(usuario_id=session["user_id"])
        pagos_query = Pago.query.filter_by(usuario_id=session["user_id"])
        
    ventas = ventas_query.order_by(Venta.fecha.desc()).all()
    
    total_subido = sum(v.monto_financiado for v in ventas)
    total_pagado = sum(p.monto for p in pagos_query.all())
    
    return render_template("dashboard.html", ventas=ventas, total_subido=total_subido, 
                           total_pagado=total_pagado, vendedores=vendedores, filtro_vendedor=filtro_vendedor)

@app.route("/venta/nueva", methods=["GET", "POST"])
def nueva_venta():
    if "user_id" not in session: return redirect(url_for("login"))
    
    if request.method == "POST":
        try:
            nueva = Venta(
                nombre_cliente=request.form.get("nombre_cliente"),
                tipo_doc=request.form.get("tipo_doc"),
                nro_doc=request.form.get("nro_doc"),
                correo=request.form.get("correo"),
                telefono=request.form.get("telefono"),
                direccion=request.form.get("direccion"),
                ciudad=request.form.get("ciudad"),
                monto_financiado=float(request.form.get("monto_financiado")),
                cantidad_cuotas=int(request.form.get("cantidad_cuotas")),
                producto_financiado=request.form.get("producto_financiado")),
                observaciones=request.form.get("observaciones"),
                
                # Conversión de adjuntos a Base64
                foto_documento=procesar_archivo("foto_documento"),
                acta_entrega=procesar_archivo("acta_entrega"),
                foto_cliente_producto=procesar_archivo("foto_cliente_producto"),
                captura_aprobacion=procesar_archivo("captura_aprobacion"),
                anexos_adicionales=procesar_archivo("anexos_adicionales"),
                usuario_id=session["user_id"]
            )
            db.session.add(nueva)
            db.session.commit()
            flash("Venta cargada exitosamente.", "success")
            return redirect(url_for("dashboard"))
        except Exception as e:
            db.session.rollback()
            flash(f"Error al guardar la venta: {str(e)}", "error")
            
    return render_template("nueva_venta.html")

@app.route("/admin/usuario", methods=["GET", "POST"])
def crear_usuario():
    if "user_id" not in session or session["role"] != "ADMIN": return redirect(url_for("login"))
    
    if request.method == "POST":
        username = request.form.get("username").strip()
        password = request.form.get("password").strip()
        role = request.form.get("role")
        
        if Usuario.query.filter_by(username=username).first():
            flash("El nombre de usuario ya existe.", "error")
        else:
            user = Usuario(username=username, password=generate_password_hash(password), role=role)
            db.session.add(user)
            db.session.commit()
            flash(f"Usuario {username} creado con éxito.", "success")
            return redirect(url_for("dashboard"))
            
    return render_template("nuevo_usuario.html")

@app.route("/admin/pago", methods=["GET", "POST"])
def registrar_pago():
    if "user_id" not in session or session["role"] != "ADMIN": return redirect(url_for("login"))
    
    vendedores = Usuario.query.filter_by(role="VENDEDOR").all()
    if request.method == "POST":
        pago = Pago(
            monto=float(request.form.get("monto")),
            observaciones=request.form.get("observaciones"),
            usuario_id=int(request.form.get("vendedor_id"))
        )
        db.session.add(pago)
        db.session.commit()
        flash("Pago registrado de manera exitosa.", "success")
        return redirect(url_for("dashboard"))
        
    return render_template("registrar_pago.html", vendedores=vendedores)

@app.route("/download_report", methods=["POST"])
def download_report():
    if "user_id" not in session: return redirect(url_for("login"))
    
    fecha_desde_str = request.form.get("fecha_desde")
    fecha_hasta_str = request.form.get("fecha_hasta")
    vendedor_filtro = request.form.get("vendedor_id_excel")
    
    query = Venta.query
    
    if fecha_desde_str:
        query = query.filter(Venta.fecha >= datetime.strptime(fecha_desde_str, "%Y-%m-%d"))
    if fecha_hasta_str:
        query = query.filter(Venta.fecha <= datetime.strptime(fecha_hasta_str, "%Y-%m-%d"))
        
    if session["role"] == "ADMIN":
        if vendedor_filtro:
            query = query.filter_by(usuario_id=int(vendedor_filtro))
    else:
        query = query.filter_by(usuario_id=session["user_id"])
        
    ventas = query.order_by(Venta.fecha.asc()).all()
    
    # Generación Estilizada del Libro Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"
    
    font_h = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    font_d = Font(name="Calibri", size=10)
    ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    headers = ["FECHA", "VENDEDOR", "CLIENTE", "TIPO DOC", "NRO DOC", "CORREO", "TELEFONO", "CIUDAD", "MONTO FINANCIADO", "CUOTAS", "PRODUCTO"]
    anchos = [14, 15, 25, 12, 15, 20, 15, 15, 18, 10, 20]
    
    for i, (col, ancho) in enumerate(zip(headers, anchos), start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = font_h
        c.fill = PatternFill("solid", start_color="1F2937") # Gris oscuro premium
        c.alignment = ac
        c.border = borde
        ws.column_dimensions[c.column_letter].width = ancho
        
    for fi, v in enumerate(ventas, start=2):
        fila_datos = [
            v.fecha.strftime("%d/%m/%Y"), v.vendedor.username, v.nombre_cliente, v.tipo_doc,
            v.nro_doc, v.correo, v.telefono, v.ciudad, v.monto_financiado, v.cantidad_cuotas, v.producto_financiado
        ]
        color = "F9FAFB" if fi % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(fila_datos, start=1):
            c = ws.cell(row=fi, column=ci, value=val)
            c.font = font_d
            c.fill = PatternFill("solid", start_color=color)
            c.border = borde
            c.alignment = ac if ci in [1, 4, 5, 7, 10] else al
            if ci == 9: c.number_format = '$#,##0.00'
            
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    
    return send_file(excel_stream, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"Reporte_Ventas_{datetime.now().strftime('%Y%m%d')}.xlsx")

if __name__ == "__main__":
    app.run(debug=True)