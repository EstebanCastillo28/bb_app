import os
import io
import base64
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__, template_folder=os.path.abspath('templates'))
app.secret_key = "vanti_sales_ultra_secure_key"

DATABASE_URL = os.getenv("POSTGRES_URL") or os.getenv("DATABASE_URL")
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL or "sqlite:///sistema_ventas.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)


class Usuario(db.Model):
    id       = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role     = db.Column(db.String(20), nullable=False)
    # Usuario congelado = activo False: no puede iniciar sesión, pero conserva sus datos.
    activo   = db.Column(db.Boolean, default=True, nullable=False)
    ventas   = db.relationship('Venta', backref='vendedor', lazy=True, cascade="all, delete-orphan")
    pagos    = db.relationship('Pago',  backref='vendedor', lazy=True, cascade="all, delete-orphan")


class Venta(db.Model):
    id                    = db.Column(db.Integer, primary_key=True)
    fecha                 = db.Column(db.DateTime, default=datetime.utcnow)
    estado                = db.Column(db.String(30), default="PENDIENTE")
    nombre_cliente        = db.Column(db.String(150), nullable=False)
    tipo_doc              = db.Column(db.String(50),  nullable=False)
    nro_doc               = db.Column(db.String(50),  nullable=False)
    correo                = db.Column(db.String(120))
    telefono              = db.Column(db.String(50))
    direccion             = db.Column(db.String(200))
    ciudad                = db.Column(db.String(100))
    monto_financiado      = db.Column(db.Float,   nullable=False)
    cantidad_cuotas       = db.Column(db.Integer, nullable=False)
    producto_financiado   = db.Column(db.String(150), nullable=False)
    observaciones         = db.Column(db.Text)
    foto_documento        = db.Column(db.Text)
    acta_entrega          = db.Column(db.Text)
    foto_cliente_producto = db.Column(db.Text)
    captura_aprobacion    = db.Column(db.Text)
    anexos_adicionales    = db.Column(db.Text)
    usuario_id            = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)


class Pago(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    fecha         = db.Column(db.DateTime, default=datetime.utcnow)
    monto         = db.Column(db.Float, nullable=False)
    observaciones = db.Column(db.Text)
    comprobante   = db.Column(db.Text)
    usuario_id    = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    # Enlace opcional con una venta. Nullable: los pagos existentes quedan sin enlazar.
    venta_id      = db.Column(db.Integer, db.ForeignKey('venta.id'), nullable=True)
    venta         = db.relationship('Venta', backref=db.backref('pagos_relacionados', lazy=True))


class Comentario(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    venta_id   = db.Column(db.Integer, db.ForeignKey('venta.id'), nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    texto      = db.Column(db.Text, nullable=False)
    fecha      = db.Column(db.DateTime, default=datetime.utcnow)
    venta      = db.relationship('Venta', backref=db.backref('comentarios', lazy=True, cascade="all, delete-orphan"))
    autor      = db.relationship('Usuario')


class CambioVenta(db.Model):
    """Registro de auditoría: cada cambio a una venta y qué usuario lo hizo."""
    id             = db.Column(db.Integer, primary_key=True)
    venta_id       = db.Column(db.Integer, db.ForeignKey('venta.id'), nullable=False)
    usuario_id     = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    campo          = db.Column(db.String(60), nullable=False)
    valor_anterior = db.Column(db.Text)
    valor_nuevo    = db.Column(db.Text)
    fecha          = db.Column(db.DateTime, default=datetime.utcnow)
    venta          = db.relationship('Venta', backref=db.backref('historial', lazy=True, cascade="all, delete-orphan"))
    autor          = db.relationship('Usuario')


def init_db():
    """Inicializa la BD de forma SEGURA: solo agrega columnas o el usuario admin
    si faltan. Nunca borra ni modifica datos ya almacenados (los ALTER TABLE son
    aditivos y fallan silenciosamente si la columna ya existe)."""
    with app.app_context():
        db.create_all()
        for ddl in (
            "ALTER TABLE venta ADD COLUMN estado VARCHAR(30) DEFAULT 'PENDIENTE'",
            "ALTER TABLE pago ADD COLUMN comprobante TEXT",
            "ALTER TABLE pago ADD COLUMN venta_id INTEGER",
            "ALTER TABLE usuario ADD COLUMN activo BOOLEAN DEFAULT true",
        ):
            try:
                db.session.execute(db.text(ddl))
                db.session.commit()
            except Exception:
                db.session.rollback()
        # Asegura que los usuarios existentes queden activos tras agregar la columna.
        try:
            db.session.execute(db.text("UPDATE usuario SET activo = true WHERE activo IS NULL"))
            db.session.commit()
        except Exception:
            db.session.rollback()
        if not Usuario.query.filter_by(username="admin").first():
            db.session.add(Usuario(
                username="admin",
                password=generate_password_hash("admin123"),
                role="ADMIN"
            ))
            db.session.commit()


# En Vercel, este bloque corre en cada arranque en frío. Si la base de datos no
# responde en ese instante, NO debe tumbar toda la función serverless
# (eso es lo que producía el error FUNCTION_INVOCATION_FAILED). Lo protegemos:
try:
    init_db()
except Exception as e:
    print(f"[init_db] No se pudo inicializar la BD al arrancar: {e}")


def procesar_archivo(file_field):
    file = request.files.get(file_field)
    if file and file.filename != '':
        return base64.b64encode(file.read()).decode('utf-8')
    return ""


# Campos que un usuario puede editar de una venta (además se auditan los cambios).
CAMPOS_EDITABLES = [
    ("nombre_cliente",      "Nombre del cliente",   "text"),
    ("tipo_doc",            "Tipo de documento",    "text"),
    ("nro_doc",             "Número de documento",  "text"),
    ("correo",              "Correo",               "text"),
    ("telefono",            "Teléfono",             "text"),
    ("direccion",           "Dirección",            "text"),
    ("ciudad",              "Ciudad",               "text"),
    ("monto_financiado",    "Monto financiado",     "number"),
    ("cantidad_cuotas",     "Cantidad de cuotas",   "number"),
    ("producto_financiado", "Producto financiado",  "text"),
    ("observaciones",       "Observaciones",        "text"),
]


def buscar_venta_duplicada(nro_doc, monto, excluir_id=None):
    """Devuelve una venta duplicada, o None. Duplicado = MISMA cédula (nro_doc)
    Y MISMO monto financiado. Así un mismo cliente puede tener varias ventas
    (con montos distintos), pero se evita cargar dos veces la misma venta.
    Se puede excluir una venta (útil al editar la venta actual)."""
    if not nro_doc:
        return None
    q = Venta.query.filter_by(nro_doc=nro_doc, monto_financiado=monto)
    if excluir_id is not None:
        q = q.filter(Venta.id != excluir_id)
    return q.first()


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username").strip()
        password = request.form.get("password").strip()
        user = Usuario.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            if user.activo is False:
                flash("Tu cuenta está congelada. Contacta al administrador.", "error")
            else:
                session["user_id"]  = user.id
                session["username"] = user.username
                session["role"]     = user.role
                return redirect(url_for("dashboard"))
        else:
            flash("Credenciales incorrectas.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
def dashboard():
    if "user_id" not in session:
        return redirect(url_for("login"))
    vendedores      = Usuario.query.filter_by(role="VENDEDOR").all() if session["role"] == "ADMIN" else []
    filtro_vendedor = request.args.get("vendedorid")
    if session["role"] == "ADMIN":
        if filtro_vendedor:
            ventas_query = Venta.query.filter_by(usuario_id=filtro_vendedor)
            pagos_query  = Pago.query.filter_by(usuario_id=filtro_vendedor)
        else:
            ventas_query = Venta.query
            pagos_query  = Pago.query
    else:
        ventas_query = Venta.query.filter_by(usuario_id=session["user_id"])
        pagos_query  = Pago.query.filter_by(usuario_id=session["user_id"])
    ventas       = ventas_query.order_by(Venta.fecha.desc()).all()
    pagos        = pagos_query.order_by(Pago.fecha.desc()).all()
    total_subido = sum(v.monto_financiado for v in ventas if v.estado != "NO INGRESO")
    total_pagado = sum(p.monto for p in pagos)
    return render_template("dashboard.html",
                           ventas=ventas, pagos=pagos,
                           total_subido=total_subido, total_pagado=total_pagado,
                           vendedores=vendedores, filtro_vendedor=filtro_vendedor)


@app.route("/venta/nueva", methods=["GET", "POST"])
def nueva_venta():
    if "user_id" not in session:
        return redirect(url_for("login"))
    if request.method == "POST":
        try:
            nro_doc = (request.form.get("nro_doc") or "").strip()
            monto   = float(request.form.get("monto_financiado"))
            duplicada = buscar_venta_duplicada(nro_doc, monto)
            if duplicada:
                flash(
                    f"No se guardó: ya existe una venta de «{duplicada.nombre_cliente}» con la "
                    f"cédula {nro_doc} por el mismo monto (${'{:,.2f}'.format(monto)}), "
                    f"cargada el {duplicada.fecha.strftime('%d/%m/%Y')}. "
                    f"Si es otra compra distinta, cambia el monto financiado.",
                    "error"
                )
                return render_template("nueva_venta.html", datos=request.form)
            nueva = Venta(
                nombre_cliente        = request.form.get("nombre_cliente"),
                tipo_doc              = request.form.get("tipo_doc"),
                nro_doc               = request.form.get("nro_doc"),
                correo                = request.form.get("correo"),
                telefono              = request.form.get("telefono"),
                direccion             = request.form.get("direccion"),
                ciudad                = request.form.get("ciudad"),
                monto_financiado      = monto,
                cantidad_cuotas       = int(request.form.get("cantidad_cuotas")),
                producto_financiado   = request.form.get("producto_financiado"),
                observaciones         = request.form.get("observaciones"),
                estado                = "PENDIENTE",
                foto_documento        = procesar_archivo("foto_documento"),
                acta_entrega          = procesar_archivo("acta_entrega"),
                foto_cliente_producto = procesar_archivo("foto_cliente_producto"),
                captura_aprobacion    = procesar_archivo("captura_aprobacion"),
                anexos_adicionales    = procesar_archivo("anexos_adicionales"),
                usuario_id            = session["user_id"]
            )
            db.session.add(nueva)
            db.session.commit()
            flash("Venta cargada exitosamente.", "success")
            return redirect(url_for("dashboard"))
        except Exception as e:
            db.session.rollback()
            flash(f"Error al guardar la venta: {str(e)}", "error")
            return render_template("nueva_venta.html", datos=request.form)
    return render_template("nueva_venta.html", datos={})


@app.route("/venta/<int:id>")
def detalle_venta(id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    venta = Venta.query.get_or_404(id)
    if session["role"] != "ADMIN" and venta.usuario_id != session["user_id"]:
        flash("Acceso denegado.", "error")
        return redirect(url_for("dashboard"))
    return render_template("detalle_venta.html", venta=venta)


@app.route("/venta/<int:id>/editar", methods=["GET", "POST"])
def editar_venta(id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    venta = Venta.query.get_or_404(id)
    # Permiso: el admin edita cualquiera; el vendedor solo las suyas.
    if session["role"] != "ADMIN" and venta.usuario_id != session["user_id"]:
        flash("No tienes autorización para editar esta venta.", "error")
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        try:
            nro_doc = (request.form.get("nro_doc") or "").strip()
            monto   = float(request.form.get("monto_financiado"))
            duplicada = buscar_venta_duplicada(nro_doc, monto, excluir_id=venta.id)
            if duplicada:
                flash(
                    f"No se guardó: ya existe otra venta de «{duplicada.nombre_cliente}» con la "
                    f"cédula {nro_doc} por el mismo monto (${'{:,.2f}'.format(monto)}). "
                    f"Si es una compra distinta, usa un monto financiado diferente.",
                    "error"
                )
                return render_template("editar_venta.html", venta=venta, campos=CAMPOS_EDITABLES)

            n_cambios = 0
            for campo, etiqueta, tipo in CAMPOS_EDITABLES:
                anterior = getattr(venta, campo)
                crudo    = request.form.get(campo)
                if tipo == "number":
                    nuevo  = float(crudo) if campo == "monto_financiado" else int(crudo)
                    cambio = float(anterior or 0) != float(nuevo)
                else:
                    nuevo  = (crudo or "").strip()
                    cambio = (anterior or "") != nuevo
                if cambio:
                    db.session.add(CambioVenta(
                        venta_id       = venta.id,
                        usuario_id     = session["user_id"],
                        campo          = etiqueta,
                        valor_anterior = "" if anterior is None else str(anterior),
                        valor_nuevo    = "" if nuevo is None else str(nuevo),
                    ))
                    setattr(venta, campo, nuevo)
                    n_cambios += 1
            db.session.commit()
            flash(f"Venta actualizada. {n_cambios} campo(s) modificado(s)." if n_cambios
                  else "No se detectaron cambios.", "success")
            return redirect(url_for("detalle_venta", id=venta.id))
        except Exception as e:
            db.session.rollback()
            flash(f"Error al actualizar la venta: {str(e)}", "error")
    return render_template("editar_venta.html", venta=venta, campos=CAMPOS_EDITABLES)


@app.route("/venta/<int:id>/comentario", methods=["POST"])
def agregar_comentario(id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    venta = Venta.query.get_or_404(id)
    if session["role"] != "ADMIN" and venta.usuario_id != session["user_id"]:
        flash("No tienes autorización para comentar esta venta.", "error")
        return redirect(url_for("dashboard"))
    texto = (request.form.get("texto") or "").strip()
    if texto:
        db.session.add(Comentario(
            venta_id=venta.id, usuario_id=session["user_id"], texto=texto
        ))
        db.session.commit()
        flash("Comentario agregado.", "success")
    return redirect(url_for("detalle_venta", id=venta.id))


@app.route("/pago/<int:id>")
def detalle_pago(id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    pago = Pago.query.get_or_404(id)
    if session["role"] != "ADMIN" and pago.usuario_id != session["user_id"]:
        flash("Acceso denegado a este comprobante.", "error")
        return redirect(url_for("dashboard"))
    return render_template("detalle_pago.html", pago=pago)


@app.route("/venta/<int:id>/estado", methods=["POST"])
def cambiar_estado(id):
    if "user_id" not in session or session["role"] != "ADMIN":
        return redirect(url_for("login"))
    venta = Venta.query.get_or_404(id)
    nuevo_estado = request.form.get("estado")
    if nuevo_estado in ["PENDIENTE", "REVISADO", "VENTA PAGA", "NO INGRESO"]:
        venta.estado = nuevo_estado
        db.session.commit()
        flash(f"Estado actualizado a {nuevo_estado}.", "success")
    return redirect(url_for("dashboard"))


@app.route("/venta/<int:id>/eliminar", methods=["POST"])
def eliminar_venta(id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    venta = Venta.query.get_or_404(id)
    if session["role"] != "ADMIN" and venta.usuario_id != session["user_id"]:
        flash("No tienes autorización para eliminar esta venta.", "error")
        return redirect(url_for("dashboard"))
    try:
        db.session.delete(venta)
        db.session.commit()
        flash("Venta eliminada permanentemente.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error al eliminar: {str(e)}", "error")
    return redirect(url_for("dashboard"))


@app.route("/pago/<int:id>/editar", methods=["GET", "POST"])
def editar_pago(id):
    if "user_id" not in session or session["role"] != "ADMIN":
        return redirect(url_for("login"))
    pago       = Pago.query.get_or_404(id)
    vendedores = Usuario.query.filter_by(role="VENDEDOR").all()
    ventas     = Venta.query.order_by(Venta.fecha.desc()).all()
    if request.method == "POST":
        try:
            venta_id_raw       = request.form.get("venta_id")
            pago.monto         = float(request.form.get("monto"))
            pago.observaciones = request.form.get("observaciones")
            pago.usuario_id    = int(request.form.get("vendedor_id"))
            pago.venta_id      = int(venta_id_raw) if venta_id_raw else None
            # El comprobante solo se reemplaza si se sube un archivo nuevo.
            nuevo_comprobante = procesar_archivo("comprobante")
            if nuevo_comprobante:
                pago.comprobante = nuevo_comprobante
            # Si queda enlazado a una venta, esa venta pasa a VENTA PAGA.
            if pago.venta_id:
                venta_enlazada = Venta.query.get(pago.venta_id)
                if venta_enlazada:
                    venta_enlazada.estado = "VENTA PAGA"
            db.session.commit()
            flash("Pago actualizado correctamente.", "success")
            return redirect(url_for("detalle_pago", id=pago.id))
        except Exception as e:
            db.session.rollback()
            flash(f"Error al actualizar el pago: {str(e)}", "error")
    return render_template("editar_pago.html", pago=pago, vendedores=vendedores, ventas=ventas)


@app.route("/pago/<int:id>/eliminar", methods=["POST"])
def eliminar_pago(id):
    if "user_id" not in session or session["role"] != "ADMIN":
        return redirect(url_for("login"))
    pago = Pago.query.get_or_404(id)
    try:
        db.session.delete(pago)
        db.session.commit()
        flash("Registro de pago eliminado con éxito.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error al eliminar el pago: {str(e)}", "error")
    return redirect(url_for("dashboard"))


@app.route("/admin/usuarios", methods=["GET", "POST"])
def admin_usuarios():
    if "user_id" not in session or session["role"] != "ADMIN":
        return redirect(url_for("login"))
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()
        role     = request.form.get("role")
        if not username or not password:
            flash("Usuario y contraseña son obligatorios.", "error")
        elif Usuario.query.filter_by(username=username).first():
            flash("El nombre de usuario ya existe.", "error")
        else:
            db.session.add(Usuario(
                username=username,
                password=generate_password_hash(password),
                role=role if role in ("ADMIN", "VENDEDOR") else "VENDEDOR",
                activo=True
            ))
            db.session.commit()
            flash(f"Usuario @{username} creado con éxito.", "success")
        return redirect(url_for("admin_usuarios"))
    usuarios = Usuario.query.order_by(Usuario.role, Usuario.username).all()
    return render_template("usuarios.html", usuarios=usuarios)


# Compatibilidad: la ruta antigua ahora lleva al nuevo panel.
@app.route("/admin/usuario")
def crear_usuario():
    return redirect(url_for("admin_usuarios"))


@app.route("/admin/usuario/<int:id>/editar", methods=["POST"])
def editar_usuario(id):
    if "user_id" not in session or session["role"] != "ADMIN":
        return redirect(url_for("login"))
    u = Usuario.query.get_or_404(id)
    nuevo_username = (request.form.get("username") or "").strip()
    nuevo_role     = request.form.get("role")
    nueva_pass     = (request.form.get("password") or "").strip()
    if nuevo_username and nuevo_username != u.username:
        if Usuario.query.filter(Usuario.username == nuevo_username, Usuario.id != u.id).first():
            flash("Ese nombre de usuario ya está en uso.", "error")
            return redirect(url_for("admin_usuarios"))
        u.username = nuevo_username
    if nuevo_role in ("ADMIN", "VENDEDOR"):
        u.role = nuevo_role
    if nueva_pass:
        u.password = generate_password_hash(nueva_pass)
    db.session.commit()
    flash(f"Usuario @{u.username} actualizado.", "success")
    return redirect(url_for("admin_usuarios"))


@app.route("/admin/usuario/<int:id>/congelar", methods=["POST"])
def congelar_usuario(id):
    if "user_id" not in session or session["role"] != "ADMIN":
        return redirect(url_for("login"))
    u = Usuario.query.get_or_404(id)
    if u.id == session["user_id"]:
        flash("No puedes congelar tu propia cuenta.", "error")
        return redirect(url_for("admin_usuarios"))
    u.activo = not bool(u.activo)
    db.session.commit()
    flash(f"Usuario @{u.username} {'reactivado' if u.activo else 'congelado'}.", "success")
    return redirect(url_for("admin_usuarios"))


@app.route("/admin/usuario/<int:id>/eliminar", methods=["POST"])
def eliminar_usuario(id):
    if "user_id" not in session or session["role"] != "ADMIN":
        return redirect(url_for("login"))
    u = Usuario.query.get_or_404(id)
    if u.id == session["user_id"]:
        flash("No puedes eliminar tu propia cuenta.", "error")
        return redirect(url_for("admin_usuarios"))
    n_ventas, n_pagos = len(u.ventas), len(u.pagos)
    if n_ventas or n_pagos:
        flash(
            f"No se eliminó: @{u.username} tiene {n_ventas} venta(s) y {n_pagos} pago(s) asociados. "
            f"Congélalo para bloquear su acceso sin borrar información.",
            "error"
        )
        return redirect(url_for("admin_usuarios"))
    db.session.delete(u)
    db.session.commit()
    flash(f"Usuario @{u.username} eliminado.", "success")
    return redirect(url_for("admin_usuarios"))


@app.route("/admin/pago", methods=["GET", "POST"])
def registrar_pago():
    if "user_id" not in session or session["role"] != "ADMIN":
        return redirect(url_for("login"))
    vendedores = Usuario.query.filter_by(role="VENDEDOR").all()
    ventas     = Venta.query.order_by(Venta.fecha.desc()).all()
    if request.method == "POST":
        try:
            venta_id_raw = request.form.get("venta_id")
            pago = Pago(
                monto         = float(request.form.get("monto")),
                observaciones = request.form.get("observaciones"),
                comprobante   = procesar_archivo("comprobante"),
                usuario_id    = int(request.form.get("vendedor_id")),
                venta_id      = int(venta_id_raw) if venta_id_raw else None
            )
            db.session.add(pago)
            # Al enlazar un pago con una venta, la venta pasa automáticamente a VENTA PAGA.
            if pago.venta_id:
                venta_enlazada = Venta.query.get(pago.venta_id)
                if venta_enlazada:
                    venta_enlazada.estado = "VENTA PAGA"
            db.session.commit()
            flash("Pago registrado de manera exitosa.", "success")
            return redirect(url_for("dashboard"))
        except Exception as e:
            db.session.rollback()
            flash(f"Error al registrar el pago: {str(e)}", "error")
    return render_template("registrar_pago.html", vendedores=vendedores, ventas=ventas)


@app.route("/download-report", methods=["POST"])
def download_report():
    if "user_id" not in session:
        return redirect(url_for("login"))
    fecha_desde_str = request.form.get("fechadesde")
    fecha_hasta_str = request.form.get("fechahasta")
    vendedor_filtro = request.form.get("vendedoridexcel")

    query = Venta.query
    if fecha_desde_str:
        query = query.filter(Venta.fecha >= datetime.strptime(fecha_desde_str, "%Y-%m-%d"))
    if fecha_hasta_str:
        limit_hasta = datetime.strptime(fecha_hasta_str, "%Y-%m-%d") + timedelta(days=1)
        query = query.filter(Venta.fecha < limit_hasta)
    if session["role"] == "ADMIN":
        if vendedor_filtro:
            query = query.filter_by(usuario_id=int(vendedor_filtro))
    else:
        query = query.filter_by(usuario_id=session["user_id"])

    ventas = query.order_by(Venta.fecha.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    font_h = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    font_d = Font(name="Calibri", size=10)
    ac    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    headers = ["FECHA","ESTADO","VENDEDOR","CLIENTE","TIPO DOC","NRO DOC","MONTO FINANCIADO","CUOTAS","PRODUCTO"]
    anchos  = [14, 15, 15, 25, 12, 15, 18, 10, 20]

    for i, (col, ancho) in enumerate(zip(headers, anchos), start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font      = font_h
        c.fill      = PatternFill("solid", start_color="1F2937")
        c.alignment = ac
        c.border    = borde
        ws.column_dimensions[c.column_letter].width = ancho

    for fi, v in enumerate(ventas, start=2):
        fila_datos = [
            v.fecha.strftime("%d/%m/%Y"), v.estado, v.vendedor.username,
            v.nombre_cliente, v.tipo_doc, v.nro_doc,
            v.monto_financiado, v.cantidad_cuotas, v.producto_financiado
        ]
        color = "F9FAFB" if fi % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(fila_datos, start=1):
            c = ws.cell(row=fi, column=ci, value=val)
            c.font      = font_d
            c.fill      = PatternFill("solid", start_color=color)
            c.border    = borde
            c.alignment = ac if ci in [1, 2, 5, 6, 8] else al
            if ci == 7:
                c.number_format = '$#,##0.00'

    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return send_file(
        excel_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Reporte_Ventas_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )
